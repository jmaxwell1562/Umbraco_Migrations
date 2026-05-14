import asyncio
import requests
import re
import xml.etree.ElementTree as ET
from playwright.async_api import async_playwright
from difflib import SequenceMatcher
import csv
import os
import argparse
import shutil
from datetime import datetime
from urllib.parse import urlparse
from html import escape
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# === 1. CONFIGURATION & ARGS ===
parser = argparse.ArgumentParser(description="WSU Migration Auditor - Advanced Scoring")
parser.add_argument("--site", help="Site Name", default="Site")
parser.add_argument("--source", help="Source URL (Prod)", default="")
parser.add_argument("--test_url", help="Full test URL, e.g. https://w2-testing.asis.wsu.edu/cougarcard/", default="")
parser.add_argument("--test_slug", help="Deprecated: W2 folder name OR full URL", default="/")
parser.add_argument(
    "--test_scope",
    choices=["ask", "single", "instance"],
    default="ask",
    help="single: only test_url, instance: include sibling instance sites, ask: prompt",
)
parser.add_argument(
    "--test_allowlist",
    default="",
    help="Optional comma-separated sibling site labels or full URLs to include (e.g., cougarcard,handbook)",
)
parser.add_argument(
    "--test_allowlist_file",
    default="",
    help="Optional text file path with one site label or URL per line for instance scope filtering",
)
parser.add_argument(
    "--redirect_override_paths",
    default="",
    help="Optional comma-separated paths to treat as intentional redirect/replacement when source is 200 and test is 404",
)
parser.add_argument("--max_tabs", help="Parallel pages", type=int, default=5)
parser.add_argument("--max_paths", help="Limit number of sitemap paths for quick smoke runs (0 = all)", type=int, default=0)
args = parser.parse_args()

SITE_NAME = args.site
SOURCE_BASE = args.source.rstrip("/")
SAFE_SITE_NAME = re.sub(r"[^A-Za-z0-9_-]+", "_", SITE_NAME).strip("_") or "Site"

# DOMAIN-AWARE TEST_BASE: Prefer explicit full URL, fallback to legacy slug behavior.
if args.test_url:
    TEST_BASE = args.test_url.rstrip("/")
elif args.test_slug.startswith("http"):
    TEST_BASE = args.test_slug.rstrip("/")
else:
    TEST_BASE = f"https://w2-testing.asis.wsu.edu/{args.test_slug.strip('/')}".rstrip("/")

PATHS_FILE = "paths.txt"


def normalize_path(path):
    cleaned = (path or "").strip()
    if not cleaned:
        return "/"
    return "/" + cleaned.strip("/") + "/"


redirect_override_paths = {normalize_path("/nutrition/net-nutrition/")}
if args.redirect_override_paths.strip():
    redirect_override_paths.update(
        normalize_path(item)
        for item in args.redirect_override_paths.split(",")
        if item.strip()
    )

# --- REVISED SCORING THRESHOLDS ---
PASS_THRESHOLD = 0.95
SOFT_PASS_THRESHOLD = 0.90

# Setup folder structure
TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
RUN_DATE = datetime.now().strftime("%Y%m%d")
BASE_DIR = f"Audit_{SITE_NAME}_{TIMESTAMP}"
os.makedirs(BASE_DIR, exist_ok=True)

semaphore = asyncio.Semaphore(args.max_tabs)


def resolve_test_scope(selection):
    if selection in {"single", "instance"}:
        return selection
    answer = input("Audit only this test URL? (y = single URL, n = include sibling instance sites): ").strip().lower()
    return "single" if answer in {"", "y", "yes"} else "instance"


def normalize_site_base(url):
    return (url or "").rstrip("/")


def source_site_label(source_base):
    parsed = urlparse(source_base or "")
    host = (parsed.hostname or "").lower()
    if not host:
        return ""
    return host.split(".")[0]


def should_prefix_test_base_path(base_path, source_base):
    """Only prefix test base path when it appears to be the site root slug."""
    cleaned = (base_path or "").strip("/").lower()
    if not cleaned:
        return False

    # Treat multi-segment paths as section/deep links, not site-root slugs.
    if "/" in cleaned:
        return False

    return cleaned == source_site_label(source_base)


def join_test_url(base, path):
    base = normalize_site_base(base)
    normalized_path = "/" + (path or "").lstrip("/")

    parsed_base = urlparse(base)
    base_path = "/" + parsed_base.path.strip("/") if parsed_base.path.strip("/") else ""

    # If test_url path looks like a section (not the source site slug),
    # join against host root so whole-site audits are not accidentally scoped.
    if base_path and not should_prefix_test_base_path(base_path, SOURCE_BASE):
        host_root = f"{parsed_base.scheme}://{parsed_base.netloc}".rstrip("/")
        return f"{host_root}{normalized_path}"

    # Avoid duplicate segments when base already includes the same leading section,
    # e.g. base=/about-urec and path=/about-urec/contact/.
    if base_path and normalized_path.startswith(base_path + "/"):
        normalized_path = normalized_path[len(base_path):]
    elif base_path and normalized_path.rstrip("/") == base_path:
        normalized_path = "/"

    if not normalized_path.startswith("/"):
        normalized_path = "/" + normalized_path

    return f"{base}{normalized_path}"


def test_site_label(test_base):
    parsed = urlparse(test_base)
    cleaned = parsed.path.strip("/")
    return cleaned if cleaned else "root"


def discover_instance_test_bases(seed_base):
    parsed = urlparse(seed_base)
    if not parsed.scheme or not parsed.netloc:
        return [normalize_site_base(seed_base)]

    host_root = f"{parsed.scheme}://{parsed.netloc}"
    discovered = []

    seed_clean = normalize_site_base(seed_base)
    if seed_clean:
        discovered.append(seed_clean)

    try:
        response = requests.get(host_root, timeout=15, **request_kwargs_for_url(host_root))
        links = re.findall(r'href=["\']([^"\']+)["\']', response.text, flags=re.IGNORECASE)
        top_level_paths = set()
        for link in links:
            if not link.startswith("/"):
                continue
            path_only = link.split("?", 1)[0].split("#", 1)[0]
            parts = [p for p in path_only.strip("/").split("/") if p]
            if len(parts) != 1:
                continue
            segment = parts[0]
            if "." in segment:
                continue
            top_level_paths.add(segment)

        for segment in sorted(top_level_paths):
            candidate = f"{host_root}/{segment}"
            if candidate not in discovered:
                discovered.append(candidate)
    except Exception as e:
        print(f"[WARN] Instance site discovery failed for {host_root}: {str(e)[:120]}")

    return discovered or [seed_clean]


def parse_test_allowlist(raw_value):
    if not raw_value:
        return set(), set()

    labels = set()
    urls = set()
    for item in raw_value.split(","):
        token = item.strip().rstrip("/")
        if not token:
            continue
        if token.startswith("http://") or token.startswith("https://"):
            urls.add(token.lower())
        else:
            labels.add(token.strip("/").lower())
    return labels, urls


def load_test_allowlist_file(file_path):
    labels = set()
    urls = set()
    if not file_path:
        return labels, urls

    try:
        with open(file_path, "r", encoding="utf-8") as f:
            for line in f:
                cleaned = line.strip()
                if not cleaned or cleaned.startswith("#"):
                    continue

                # Permit comma-separated entries and one-per-line formats.
                for item in cleaned.split(","):
                    token = item.strip().rstrip("/")
                    if not token:
                        continue
                    if token.startswith("http://") or token.startswith("https://"):
                        urls.add(token.lower())
                    else:
                        labels.add(token.strip("/").lower())
    except Exception as e:
        print(f"[WARN] Failed to read test_allowlist_file '{file_path}': {str(e)[:120]}")

    return labels, urls


def build_test_allowlist(raw_value, allowlist_file):
    labels, urls = parse_test_allowlist(raw_value)
    file_labels, file_urls = load_test_allowlist_file(allowlist_file)
    labels.update(file_labels)
    urls.update(file_urls)
    return labels, urls


def filter_test_bases_by_allowlist(test_bases, labels, urls):
    if not labels and not urls:
        return test_bases

    filtered = []
    for base in test_bases:
        normalized = normalize_site_base(base)
        normalized_lc = normalized.lower()
        label = test_site_label(normalized).lower()
        if normalized_lc in urls or label in labels:
            filtered.append(normalized)

    return filtered


def request_kwargs_for_url(url):
    parsed = urlparse(url or "")
    host = (parsed.hostname or "").lower()
    if host in {"localhost", "127.0.0.1", "::1"}:
        return {"verify": False}
    return {}


def is_localhost_url(url):
    parsed = urlparse(url or "")
    host = (parsed.hostname or "").lower()
    return host in {"localhost", "127.0.0.1", "::1"}


def probe_url(url):
    """Fetch URL once without following redirects to record HTTP status and redirect flag."""
    try:
        response = requests.get(url, allow_redirects=False, timeout=20, **request_kwargs_for_url(url))
        is_redirect = response.is_redirect or response.is_permanent_redirect
        return response.status_code, is_redirect
    except Exception as e:
        print(f"[WARN] Probe failed for {url}: {str(e)[:120]}")
        return None, False


def maybe_correct_test_base(source_base, test_base, paths):
    """Detect likely w2/wdev3 host mismatch and auto-correct test_base when evidence is strong."""
    # Skip host-mismatch detection if no source base provided
    if not source_base:
        return test_base
    
    parsed = urlparse(test_base)
    host = (parsed.netloc or "").lower()

    host_map = {
        "wdev3-testing.asis.wsu.edu": "w2-testing.asis.wsu.edu",
        "w2-testing.asis.wsu.edu": "wdev3-testing.asis.wsu.edu",
    }
    if host not in host_map:
        return test_base

    candidate_paths = [p for p in paths if p and p != "/"]
    sample_paths = candidate_paths[:12]
    if not sample_paths:
        return test_base

    suspect_paths = []
    for path in sample_paths:
        source_url = f"{source_base}{path}" if source_base else None
        source_status, _ = probe_url(source_url) if source_url else (None, False)
        test_status, _ = probe_url(f"{test_base}{path}")
        if source_status == 200 and (test_status is None or test_status >= 400):
            suspect_paths.append(path)

    if len(suspect_paths) < max(3, int(len(sample_paths) * 0.6)):
        return test_base

    alt_host = host_map[host]
    alt_path = parsed.path.rstrip("/")
    alt_base = f"{parsed.scheme}://{alt_host}{alt_path}" if alt_path else f"{parsed.scheme}://{alt_host}"
    recovered = 0
    for path in suspect_paths:
        alt_status, _ = probe_url(f"{alt_base}{path}")
        if alt_status is not None and alt_status < 400:
            recovered += 1

    if recovered >= max(3, int(len(suspect_paths) * 0.6)):
        print(
            "[WARN] Detected probable test-host mismatch "
            f"({host} -> {alt_host}) based on sampled path health; auto-switching to {alt_base}"
        )
        return alt_base

    return test_base


def determine_root_cause(status, score, source_status, test_status, redirect):
    if redirect:
        return "Redirect handling required"
    if status == "FAIL":
        if source_status == 200 and (test_status is None or test_status >= 400):
            return "Migration gap: source healthy, test failing"
        if source_status is not None and source_status >= 400 and test_status is not None and test_status >= 400:
            return "Shared instability: source and test both failing"
        if test_status is not None and test_status >= 500:
            return "Test server error"
        if test_status is not None and test_status == 404:
            return "Missing route/content on test"
        return "Failing page needs manual investigation"
    if status == "REVIEW":
        return f"Content mismatch: {score:.1%} similarity"
    if status == "SOFT PASS":
        return "Minor content differences"
    if status == "PASS":
        return "High content parity"
    if status == "SKIP":
        return "Redirected URL skipped"
    return "Unexpected error during audit"


def sort_audit_results(rows):
    status_order = {
        "FAIL": 0,
        "REVIEW": 1,
        "SOFT PASS": 2,
        "PASS": 3,
        "SKIP": 4,
        "ERROR": 5,
    }

    def to_float(value):
        try:
            return float(value)
        except (TypeError, ValueError):
            return 0.0

    rows.sort(
        key=lambda row: (
            status_order.get((row.get("status") or "").strip(), 99),
            to_float(row.get("score", 0)),
            row.get("test_site") or "",
            row.get("path") or "",
        )
    )
    return rows


def section_from_path(path):
    cleaned = (path or "").strip()
    if not cleaned or cleaned == "/":
        return "home"

    parts = [p for p in cleaned.strip("/").split("/") if p]
    if not parts:
        return "home"
    return parts[0]


def normalize_signature(row):
    root_cause = (row.get("root_cause") or "").strip().lower()
    test_status = row.get("test_status")
    source_status = row.get("source_status")

    if root_cause:
        return root_cause

    if isinstance(test_status, int) and test_status >= 500:
        return "test server error"
    if isinstance(test_status, int) and test_status == 404:
        return "missing route/content on test"
    if isinstance(source_status, int) and isinstance(test_status, int):
        return f"http pattern {source_status}/{test_status}"
    return "unknown failure signature"


def cluster_failure_patterns(rows):
    failure_statuses = {"FAIL", "ERROR"}
    failures = [row for row in rows if (row.get("status") or "").strip() in failure_statuses]

    # Initialize clustering fields for every row.
    for row in rows:
        row["section"] = section_from_path(row.get("path", ""))
        row["test_site"] = row.get("test_site") or "root"
        row["failure_cluster"] = ""
        row["systemic_breakage"] = False
        row["systemic_reason"] = ""

    if not failures:
        return []

    grouped = {}
    signature_sections = {}
    for row in failures:
        test_site = row["test_site"]
        section = row["section"]
        signature = normalize_signature(row)
        key = (test_site, section, signature)
        grouped.setdefault(key, []).append(row)
        signature_sections.setdefault((test_site, signature), set()).add(section)

    summary_rows = []
    section_cluster_counts = {}
    for key, cluster_rows in sorted(grouped.items(), key=lambda item: (-len(item[1]), item[0][0], item[0][1], item[0][2])):
        test_site, section, signature = key
        section_scope = (test_site, section)
        section_cluster_counts[section_scope] = section_cluster_counts.get(section_scope, 0) + 1
        cluster_id = f"{test_site}:{section}#{section_cluster_counts[section_scope]}"

        section_failures = [r for r in failures if r["test_site"] == test_site and r["section"] == section]
        section_failure_count = len(section_failures)
        cluster_count = len(cluster_rows)
        section_ratio = (cluster_count / section_failure_count) if section_failure_count else 0.0

        signature_key = (test_site, signature)
        cross_section_count = len(signature_sections.get(signature_key, set()))
        cross_section_total = sum(
            len(grouped.get((test_site, other_section, signature), []))
            for other_section in signature_sections.get(signature_key, set())
        )

        section_level_systemic = cluster_count >= 3 and section_ratio >= 0.60
        cross_section_systemic = cross_section_count >= 3 and cross_section_total >= 6

        is_systemic = section_level_systemic or cross_section_systemic
        reason_parts = []
        if section_level_systemic:
            reason_parts.append(
                f"Section pattern: {cluster_count}/{section_failure_count} failures in '{section}' share '{signature}'"
            )
        if cross_section_systemic:
            reason_parts.append(
                f"Shared pattern: '{signature}' appears across {cross_section_count} sections ({cross_section_total} failures)"
            )
        systemic_reason = " | ".join(reason_parts)

        for row in cluster_rows:
            row["failure_cluster"] = cluster_id
            row["systemic_breakage"] = is_systemic
            row["systemic_reason"] = systemic_reason

        summary_rows.append(
            {
                "failure_cluster": cluster_id,
                "test_site": test_site,
                "section": section,
                "signature": signature,
                "failures_in_cluster": cluster_count,
                "section_failures": section_failure_count,
                "sections_with_signature": cross_section_count,
                "total_failures_with_signature": cross_section_total,
                "systemic_breakage": is_systemic,
                "systemic_reason": systemic_reason,
            }
        )

    return summary_rows


def write_cluster_summary(base_dir, safe_site_name, run_date, summary_rows):
    summary_name = f"{safe_site_name}_failure_clusters_{run_date}.csv"
    summary_path = os.path.join(base_dir, summary_name)
    with open(summary_path, 'w', newline='') as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "failure_cluster",
                "test_site",
                "section",
                "signature",
                "failures_in_cluster",
                "section_failures",
                "sections_with_signature",
                "total_failures_with_signature",
                "systemic_breakage",
                "systemic_reason",
            ],
        )
        writer.writeheader()
        writer.writerows(summary_rows)
    return summary_path


def classify_release_impact(row):
    status = (row.get("status") or "").strip()
    source_status = row.get("source_status")
    test_status = row.get("test_status")
    systemic_breakage = bool(row.get("systemic_breakage"))
    root_cause = (row.get("root_cause") or "").strip()

    blocker = False
    reasons = []

    if status == "ERROR":
        blocker = True
        reasons.append("Runtime audit error")

    if status == "FAIL":
        blocker = True
        if source_status == 200 and (test_status is None or test_status >= 400):
            reasons.append("Migration gap")
        else:
            reasons.append("Page failure")

    if systemic_breakage:
        blocker = True
        reasons.append("Systemic breakage cluster")

    if status == "SKIP" and "Redirect handling required" in root_cause:
        reasons.append("Redirect detected")

    blocker_reason = " | ".join(dict.fromkeys(reasons))
    release_impact = "blocker" if blocker else "non-blocker"
    return blocker, blocker_reason, release_impact


def classify_change_bucket(row):
    status = (row.get("status") or "").strip()
    try:
        score = float(row.get("score", 0) or 0)
    except (TypeError, ValueError):
        score = 0.0

    if status in {"PASS", "SOFT PASS"}:
        return "none"
    if status == "REDIRECT":
        return "redirect"
    if status in {"FAIL", "ERROR"}:
        return "major"
    if score >= 0.88:
        return "minor"
    if score >= 0.70:
        return "moderate"
    return "major"


def derive_fix_code(row):
    status = (row.get("status") or "").strip()
    source_status = row.get("source_status")
    test_status = row.get("test_status")
    root_cause = (row.get("root_cause") or "").strip().lower()

    if status == "REDIRECT":
        return "INTENTIONAL_REDIRECT"
    if status == "ERROR":
        return "AUDIT_ERROR"
    if status == "FAIL":
        if source_status == 200 and (test_status is None or test_status >= 400):
            return "MISSING_PAGE"
        if isinstance(test_status, int) and test_status >= 500:
            return "TEST_SERVER_ERROR"
        if isinstance(source_status, int) and source_status >= 400 and isinstance(test_status, int) and test_status >= 400:
            return "SOURCE_ISSUE"
        return "PAGE_FAILURE"
    if status in {"REVIEW", "SOFT PASS"}:
        return "CONTENT_DRIFT"
    if status == "PASS":
        return "PARITY_OK"
    if "redirect" in root_cause:
        return "INTENTIONAL_REDIRECT"
    return "MANUAL_REVIEW"


def derive_owner(section):
    section_key = (section or "").strip().lower()
    owner_map = {
        "search": "platform-search",
        "searchpage": "platform-search",
        "home": "web-content",
        "events": "web-content",
        "about-urec": "urec-content",
        "programs-classes": "urec-programs",
        "locations-facilities": "urec-facilities",
        "memberships-lockers": "urec-memberships",
    }
    return owner_map.get(section_key, "web-content")


def build_release_readiness(rows):
    sections = {}

    for row in rows:
        test_site = row.get("test_site") or "root"
        row["test_site"] = test_site
        section = row.get("section") or section_from_path(row.get("path", ""))
        row["section"] = section

        blocker, blocker_reason, release_impact = classify_release_impact(row)
        row["blocker"] = blocker
        row["blocker_reason"] = blocker_reason
        row["release_impact"] = release_impact
        row["change_bucket"] = classify_change_bucket(row)
        row["fix_code"] = derive_fix_code(row)
        row["owner"] = derive_owner(section)

        scope_key = (test_site, section)
        section_bucket = sections.setdefault(
            scope_key,
            {
                "test_site": test_site,
                "section": section,
                "total_pages": 0,
                "blocker_count": 0,
                "non_blocker_count": 0,
                "systemic_cluster_count": 0,
                "readiness_score": 100,
                "readiness_label": "GO",
                "go_no_go": "GO",
                "summary_reason": "",
                "_clusters": set(),
                "_systemic_blockers": 0,
            },
        )

        section_bucket["total_pages"] += 1
        if blocker:
            section_bucket["blocker_count"] += 1
        else:
            section_bucket["non_blocker_count"] += 1

        if bool(row.get("systemic_breakage")) and row.get("failure_cluster"):
            section_bucket["_clusters"].add(row["failure_cluster"])
            if blocker:
                section_bucket["_systemic_blockers"] += 1

        status = (row.get("status") or "").strip()
        if status == "ERROR":
            section_bucket["readiness_score"] -= 25
        elif status == "FAIL":
            section_bucket["readiness_score"] -= 20
            if bool(row.get("systemic_breakage")):
                section_bucket["readiness_score"] -= 10
        elif status == "REVIEW":
            section_bucket["readiness_score"] -= 5
        elif status == "SOFT PASS":
            section_bucket["readiness_score"] -= 2

    summary_rows = []
    for scope_key, bucket in sections.items():
        section = bucket["section"]
        score = max(0, min(100, bucket["readiness_score"]))
        blocker_count = bucket["blocker_count"]
        systemic_blockers = bucket["_systemic_blockers"]

        if score >= 90 and blocker_count == 0:
            readiness_label = "GO"
            go_no_go = "GO"
            summary_reason = "No blockers and high parity"
        elif score >= 75 and systemic_blockers == 0:
            readiness_label = "CONDITIONAL GO"
            go_no_go = "GO"
            summary_reason = "Minor issues present; no systemic blockers"
        else:
            readiness_label = "NO GO"
            go_no_go = "NO GO"
            if systemic_blockers > 0:
                summary_reason = "Systemic blockers detected"
            elif blocker_count > 0:
                summary_reason = "Blocking failures detected"
            else:
                summary_reason = "Readiness score below threshold"

        summary_rows.append(
            {
                "test_site": bucket["test_site"],
                "section": section,
                "total_pages": bucket["total_pages"],
                "blocker_count": blocker_count,
                "non_blocker_count": bucket["non_blocker_count"],
                "systemic_cluster_count": len(bucket["_clusters"]),
                "readiness_score": score,
                "readiness_label": readiness_label,
                "go_no_go": go_no_go,
                "summary_reason": summary_reason,
            }
        )

    summary_rows.sort(key=lambda item: (item["go_no_go"] != "NO GO", item["readiness_score"], item["section"]))
    return summary_rows


def write_release_readiness_summary(base_dir, safe_site_name, run_date, summary_rows):
    summary_name = f"{safe_site_name}_release_readiness_{run_date}.csv"
    summary_path = os.path.join(base_dir, summary_name)
    with open(summary_path, 'w', newline='') as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "test_site",
                "section",
                "total_pages",
                "blocker_count",
                "non_blocker_count",
                "systemic_cluster_count",
                "readiness_score",
                "readiness_label",
                "go_no_go",
                "summary_reason",
            ],
        )
        writer.writeheader()
        writer.writerows(summary_rows)
    return summary_path


def write_accessible_html_report(base_dir, safe_site_name, run_date, result_rows, cluster_rows, readiness_rows):
        html_name = f"{safe_site_name}_audit_report_{run_date}.html"
        html_path = os.path.join(base_dir, html_name)

        status_class = {
                "PASS": "status-pass",
                "SOFT PASS": "status-soft-pass",
                "REVIEW": "status-review",
                "FAIL": "status-fail",
            "REDIRECT": "status-redirect",
                "SKIP": "status-skip",
                "ERROR": "status-error",
        }
        readiness_class = {
                "GO": "readiness-go",
                "CONDITIONAL GO": "readiness-conditional",
                "NO GO": "readiness-no-go",
        }

        def td(value):
                return escape("" if value is None else str(value))

        result_rows_html = []
        detail_rows = sorted(
            result_rows,
            key=lambda row: (
                (row.get("section") or "unassigned").lower(),
                (row.get("path") or "").lower(),
            ),
        )
        current_section = None
        for row in detail_rows:
            section_label = row.get("section") or "unassigned"
            if section_label != current_section:
                current_section = section_label
                result_rows_html.append(
                    "<tr class='section-subhead'>"
                    f"<td colspan='10'>Section: {td(section_label)}</td>"
                    "</tr>"
                )

            s_class = status_class.get((row.get("status") or "").strip(), "")
            result_rows_html.append(
                "<tr>"
                f"<td>{td(row.get('path'))}</td>"
                f"<td>{td(row.get('source_url'))}</td>"
                f"<td>{td(row.get('test_url'))}</td>"
                f"<td>{td(row.get('score'))}</td>"
                f"<td><span class='pill {s_class}'>{td(row.get('status'))}</span></td>"
                f"<td>{td(row.get('note'))}</td>"
                f"<td>{td(row.get('root_cause'))}</td>"
                f"<td>{td(row.get('change_bucket'))}</td>"
                f"<td>{td(row.get('fix_code'))}</td>"
                f"<td>{td(row.get('owner'))}</td>"
                "</tr>"
            )

        readiness_rows_html = []
        for row in readiness_rows:
                r_class = readiness_class.get((row.get("readiness_label") or "").strip(), "")
                readiness_rows_html.append(
                        "<tr>"
                        f"<td>{td(row.get('section'))}</td>"
                    f"<td>{td(row.get('test_site'))}</td>"
                        f"<td>{td(row.get('total_pages'))}</td>"
                        f"<td>{td(row.get('blocker_count'))}</td>"
                        f"<td>{td(row.get('non_blocker_count'))}</td>"
                        f"<td>{td(row.get('systemic_cluster_count'))}</td>"
                        f"<td>{td(row.get('readiness_score'))}</td>"
                        f"<td><span class='pill {r_class}'>{td(row.get('readiness_label'))}</span></td>"
                        f"<td>{td(row.get('go_no_go'))}</td>"
                        f"<td>{td(row.get('summary_reason'))}</td>"
                        "</tr>"
                )



        html_content = f"""<!doctype html>
<html lang=\"en\">
<head>
    <meta charset=\"utf-8\" />
    <meta name=\"viewport\" content=\"width=device-width, initial-scale=1\" />
    <title>{escape(safe_site_name)} Audit Report {escape(run_date)}</title>
    <style>
        :root {{
            --bg: #ffffff;
            --panel: #f5f7fa;
            --text: #121417;
            --muted: #44505c;
            --border: #c7d0d9;
            --pass-bg: #0f5132;
            --pass-fg: #ffffff;
            --soft-bg: #1f4e79;
            --soft-fg: #ffffff;
            --review-bg: #8a4b00;
            --review-fg: #ffffff;
            --fail-bg: #7a0916;
            --fail-fg: #ffffff;
            --redirect-bg: #4b5563;
            --redirect-fg: #ffffff;
            --skip-bg: #5a5f66;
            --skip-fg: #ffffff;
            --error-bg: #4b2142;
            --error-fg: #ffffff;
            --go-bg: #005f3c;
            --go-fg: #ffffff;
            --cond-bg: #6a4b00;
            --cond-fg: #ffffff;
            --nogo-bg: #7a0916;
            --nogo-fg: #ffffff;
            --blocker-bg: #7a0916;
            --blocker-fg: #ffffff;
            --nonblocker-bg: #1f4e79;
            --nonblocker-fg: #ffffff;
        }}
        body {{ margin: 0; font-family: Segoe UI, Tahoma, sans-serif; color: var(--text); background: var(--bg); }}
        main {{ padding: 20px; }}
        h1, h2 {{ margin: 0 0 12px 0; }}
        p {{ color: var(--muted); margin: 0 0 14px 0; }}
        .legend {{ display: flex; flex-wrap: wrap; gap: 8px; margin-bottom: 16px; }}
        .pill {{ display: inline-block; padding: 3px 8px; border-radius: 999px; font-size: 12px; font-weight: 700; letter-spacing: 0.2px; }}
        .status-pass {{ background: var(--pass-bg); color: var(--pass-fg); }}
        .status-soft-pass {{ background: var(--soft-bg); color: var(--soft-fg); }}
        .status-review {{ background: var(--review-bg); color: var(--review-fg); }}
        .status-fail {{ background: var(--fail-bg); color: var(--fail-fg); }}
        .status-redirect {{ background: var(--redirect-bg); color: var(--redirect-fg); }}
        .status-skip {{ background: var(--skip-bg); color: var(--skip-fg); }}
        .status-error {{ background: var(--error-bg); color: var(--error-fg); }}
        .readiness-go {{ background: var(--go-bg); color: var(--go-fg); }}
        .readiness-conditional {{ background: var(--cond-bg); color: var(--cond-fg); }}
        .readiness-no-go {{ background: var(--nogo-bg); color: var(--nogo-fg); }}
        .impact-blocker {{ background: var(--blocker-bg); color: var(--blocker-fg); }}
        .impact-non-blocker {{ background: var(--nonblocker-bg); color: var(--nonblocker-fg); }}
        .table-wrap {{ overflow-x: auto; border: 1px solid var(--border); border-radius: 8px; margin-bottom: 18px; }}
        table {{ border-collapse: collapse; width: 100%; min-width: 900px; background: #fff; }}
        th, td {{ border-bottom: 1px solid var(--border); padding: 8px 10px; font-size: 13px; text-align: left; vertical-align: top; }}
        th {{ position: sticky; top: 0; background: var(--panel); z-index: 1; }}
        .section-subhead td {{ background: #e9eef5; font-weight: 700; color: #1f2937; }}
    </style>
</head>
<body>
    <main>
        <h1>{escape(safe_site_name)} Migration Audit ({escape(run_date)})</h1>
        <p>Color-coding uses high-contrast text/background combinations to support accessible readability.</p>
        <div class=\"legend\">
            <span class=\"pill status-fail\">FAIL</span>
            <span class=\"pill status-review\">REVIEW</span>
            <span class=\"pill status-soft-pass\">SOFT PASS</span>
            <span class=\"pill status-pass\">PASS</span>
            <span class=\"pill status-redirect\">REDIRECT</span>
            <span class=\"pill status-skip\">SKIP</span>
            <span class=\"pill status-error\">ERROR</span>
            <span class=\"pill readiness-go\">GO</span>
            <span class=\"pill readiness-conditional\">CONDITIONAL GO</span>
            <span class=\"pill readiness-no-go\">NO GO</span>
            <span class=\"pill impact-blocker\">blocker</span>
            <span class=\"pill impact-non-blocker\">non-blocker</span>
        </div>

        <h2>Section Release Readiness</h2>
        <div class=\"table-wrap\">
            <table>
                <thead>
                    <tr>
                        <th>section</th><th>test_site</th><th>total_pages</th><th>blocker_count</th><th>non_blocker_count</th>
                        <th>systemic_cluster_count</th><th>readiness_score</th><th>readiness_label</th>
                        <th>go_no_go</th><th>summary_reason</th>
                    </tr>
                </thead>
                <tbody>
                    {''.join(readiness_rows_html)}
                </tbody>
            </table>
        </div>

        <h2>Detailed Rows</h2>
        <div class=\"table-wrap\">
            <table>
                <thead>
                    <tr>
                        <th>path</th><th>source_url</th><th>test_url</th><th>score</th><th>status</th><th>note</th><th>root_cause</th>
                        <th>change_bucket</th><th>fix_code</th><th>owner</th>
                    </tr>
                </thead>
                <tbody>
                    {''.join(result_rows_html)}
                </tbody>
            </table>
        </div>
    </main>
</body>
</html>
"""

        with open(html_path, "w", encoding="utf-8") as f:
                f.write(html_content)
        return html_path


def write_executive_html_report(base_dir, safe_site_name, run_date, result_rows, readiness_rows):
        html_name = f"{safe_site_name}_executive_view_{run_date}.html"
        html_path = os.path.join(base_dir, html_name)

        readiness_class = {
                "GO": "readiness-go",
                "CONDITIONAL GO": "readiness-conditional",
                "NO GO": "readiness-no-go",
        }

        blocker_rows = [row for row in result_rows if row.get("release_impact") == "blocker"]

        def td(value):
                return escape("" if value is None else str(value))

        readiness_rows_html = []
        for row in readiness_rows:
                r_class = readiness_class.get((row.get("readiness_label") or "").strip(), "")
                readiness_rows_html.append(
                        "<tr>"
                        f"<td>{td(row.get('section'))}</td>"
                    f"<td>{td(row.get('test_site'))}</td>"
                        f"<td>{td(row.get('readiness_score'))}</td>"
                        f"<td><span class='pill {r_class}'>{td(row.get('readiness_label'))}</span></td>"
                        f"<td>{td(row.get('go_no_go'))}</td>"
                        f"<td>{td(row.get('blocker_count'))}</td>"
                        f"<td>{td(row.get('systemic_cluster_count'))}</td>"
                        f"<td>{td(row.get('summary_reason'))}</td>"
                        "</tr>"
                )

        blocker_rows_html = []
        for row in blocker_rows:
                blocker_rows_html.append(
                        "<tr>"
                        f"<td>{td(row.get('section'))}</td>"
                    f"<td>{td(row.get('test_site'))}</td>"
                        f"<td>{td(row.get('path'))}</td>"
                        f"<td>{td(row.get('status'))}</td>"
                        f"<td>{td(row.get('root_cause'))}</td>"
                        f"<td>{td(row.get('blocker_reason'))}</td>"
                        f"<td>{td(row.get('systemic_breakage'))}</td>"
                        "</tr>"
                )

        total_sections = len(readiness_rows)
        no_go_sections = sum(1 for row in readiness_rows if row.get("go_no_go") == "NO GO")
        conditional_sections = sum(1 for row in readiness_rows if row.get("readiness_label") == "CONDITIONAL GO")
        go_sections = sum(1 for row in readiness_rows if row.get("readiness_label") == "GO")

        html_content = f"""<!doctype html>
<html lang=\"en\">
<head>
    <meta charset=\"utf-8\" />
    <meta name=\"viewport\" content=\"width=device-width, initial-scale=1\" />
    <title>{escape(safe_site_name)} Executive View {escape(run_date)}</title>
    <style>
        :root {{
            --bg: #ffffff;
            --panel: #f5f7fa;
            --text: #121417;
            --muted: #44505c;
            --border: #c7d0d9;
            --go-bg: #005f3c;
            --go-fg: #ffffff;
            --cond-bg: #6a4b00;
            --cond-fg: #ffffff;
            --nogo-bg: #7a0916;
            --nogo-fg: #ffffff;
            --metric-bg: #eef3f8;
        }}
        body {{ margin: 0; font-family: Segoe UI, Tahoma, sans-serif; color: var(--text); background: var(--bg); }}
        main {{ padding: 20px; max-width: 1300px; margin: 0 auto; }}
        h1, h2 {{ margin: 0 0 12px 0; }}
        p {{ color: var(--muted); margin: 0 0 14px 0; }}
        .metrics {{ display: grid; grid-template-columns: repeat(4, minmax(140px, 1fr)); gap: 10px; margin-bottom: 16px; }}
        .metric {{ background: var(--metric-bg); border: 1px solid var(--border); border-radius: 8px; padding: 10px; }}
        .metric .label {{ font-size: 12px; color: var(--muted); }}
        .metric .value {{ font-size: 22px; font-weight: 700; margin-top: 4px; }}
        .legend {{ display: flex; flex-wrap: wrap; gap: 8px; margin-bottom: 16px; }}
        .pill {{ display: inline-block; padding: 3px 8px; border-radius: 999px; font-size: 12px; font-weight: 700; letter-spacing: 0.2px; }}
        .readiness-go {{ background: var(--go-bg); color: var(--go-fg); }}
        .readiness-conditional {{ background: var(--cond-bg); color: var(--cond-fg); }}
        .readiness-no-go {{ background: var(--nogo-bg); color: var(--nogo-fg); }}
        .table-wrap {{ overflow-x: auto; border: 1px solid var(--border); border-radius: 8px; margin-bottom: 18px; }}
        table {{ border-collapse: collapse; width: 100%; min-width: 900px; background: #fff; }}
        th, td {{ border-bottom: 1px solid var(--border); padding: 8px 10px; font-size: 13px; text-align: left; vertical-align: top; }}
        th {{ position: sticky; top: 0; background: var(--panel); z-index: 1; }}
        .empty {{ padding: 16px; border: 1px solid var(--border); border-radius: 8px; background: #fbfcfd; }}
        @media (max-width: 800px) {{
            .metrics {{ grid-template-columns: 1fr 1fr; }}
        }}
    </style>
</head>
<body>
    <main>
        <h1>{escape(safe_site_name)} Executive Release View ({escape(run_date)})</h1>
        <p>Compact release summary with section go/no-go outcomes and current blockers.</p>

        <div class=\"metrics\">
            <div class=\"metric\"><div class=\"label\">Sections</div><div class=\"value\">{total_sections}</div></div>
            <div class=\"metric\"><div class=\"label\">GO</div><div class=\"value\">{go_sections}</div></div>
            <div class=\"metric\"><div class=\"label\">Conditional</div><div class=\"value\">{conditional_sections}</div></div>
            <div class=\"metric\"><div class=\"label\">NO GO</div><div class=\"value\">{no_go_sections}</div></div>
        </div>

        <div class=\"legend\">
            <span class=\"pill readiness-go\">GO</span>
            <span class=\"pill readiness-conditional\">CONDITIONAL GO</span>
            <span class=\"pill readiness-no-go\">NO GO</span>
        </div>

        <h2>Section Decisions</h2>
        <div class=\"table-wrap\">
            <table>
                <thead>
                    <tr>
                        <th>section</th><th>test_site</th><th>score</th><th>readiness_label</th><th>go_no_go</th>
                        <th>blockers</th><th>systemic_clusters</th><th>summary_reason</th>
                    </tr>
                </thead>
                <tbody>
                    {''.join(readiness_rows_html)}
                </tbody>
            </table>
        </div>

        <h2>Current Blockers</h2>
        {"<div class='table-wrap'><table><thead><tr><th>section</th><th>test_site</th><th>path</th><th>status</th><th>root_cause</th><th>blocker_reason</th><th>systemic_breakage</th></tr></thead><tbody>" + ''.join(blocker_rows_html) + "</tbody></table></div>" if blocker_rows else "<div class='empty'>No blockers detected in this run.</div>"}
    </main>
</body>
</html>
"""

        with open(html_path, "w", encoding="utf-8") as f:
                f.write(html_content)
        return html_path


def write_xlsx_report(base_dir, safe_site_name, run_date, result_rows, cluster_rows, readiness_rows):
    xlsx_name = f"{safe_site_name}_audit_report_{run_date}.xlsx"
    xlsx_path = os.path.join(base_dir, xlsx_name)

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)

    status_styles = {
        "PASS": (PatternFill(fill_type="solid", fgColor="0F5132"), Font(color="FFFFFF", bold=True)),
        "SOFT PASS": (PatternFill(fill_type="solid", fgColor="1F4E79"), Font(color="FFFFFF", bold=True)),
        "REVIEW": (PatternFill(fill_type="solid", fgColor="8A4B00"), Font(color="FFFFFF", bold=True)),
        "FAIL": (PatternFill(fill_type="solid", fgColor="7A0916"), Font(color="FFFFFF", bold=True)),
        "REDIRECT": (PatternFill(fill_type="solid", fgColor="4B5563"), Font(color="FFFFFF", bold=True)),
        "SKIP": (PatternFill(fill_type="solid", fgColor="5A5F66"), Font(color="FFFFFF", bold=True)),
        "ERROR": (PatternFill(fill_type="solid", fgColor="4B2142"), Font(color="FFFFFF", bold=True)),
    }
    readiness_styles = {
        "GO": (PatternFill(fill_type="solid", fgColor="005F3C"), Font(color="FFFFFF", bold=True)),
        "CONDITIONAL GO": (PatternFill(fill_type="solid", fgColor="6A4B00"), Font(color="FFFFFF", bold=True)),
        "NO GO": (PatternFill(fill_type="solid", fgColor="7A0916"), Font(color="FFFFFF", bold=True)),
    }
    def write_sheet(name, headers, rows, style_handler=None, section_subheads=False):
        ws = wb.create_sheet(title=name)
        ws.append(headers)

        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font

        if section_subheads:
            sorted_rows = sorted(
                rows,
                key=lambda row: (
                    (row.get("section") or "unassigned").lower(),
                    (row.get("path") or "").lower(),
                ),
            )
            current_section = None
            for row in sorted_rows:
                section_label = row.get("section") or "unassigned"
                if section_label != current_section:
                    current_section = section_label
                    ws.append([f"SECTION: {section_label}"] + [""] * (len(headers) - 1))
                    subhead_row = ws.max_row
                    ws.merge_cells(start_row=subhead_row, start_column=1, end_row=subhead_row, end_column=len(headers))
                    subhead_cell = ws.cell(row=subhead_row, column=1)
                    subhead_cell.fill = PatternFill(fill_type="solid", fgColor="E9EEF5")
                    subhead_cell.font = Font(color="1F2937", bold=True)

                ws.append([row.get(h, "") for h in headers])
                if style_handler:
                    style_handler(ws, ws.max_row, row, headers)
        else:
            for row in rows:
                ws.append([row.get(h, "") for h in headers])
                if style_handler:
                    style_handler(ws, ws.max_row, row, headers)

        ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(headers)).coordinate}"
        ws.freeze_panes = "A2"

        for col_idx, header in enumerate(headers, start=1):
            max_len = len(str(header))
            for row_idx in range(2, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(80, max_len + 2)

        return ws

    def style_detail(ws, row_idx, row_data, headers):
        status = (row_data.get("status") or "").strip()
        if status in status_styles:
            fill, font = status_styles[status]
            status_col = headers.index("status") + 1
            ws.cell(row=row_idx, column=status_col).fill = fill
            ws.cell(row=row_idx, column=status_col).font = font

    def style_cluster(ws, row_idx, row_data, headers):
        systemic = row_data.get("systemic_breakage")
        col = headers.index("systemic_breakage") + 1
        cell = ws.cell(row=row_idx, column=col)
        if str(systemic).lower() == "true":
            cell.fill = PatternFill(fill_type="solid", fgColor="7A0916")
            cell.font = Font(color="FFFFFF", bold=True)
        elif str(systemic).lower() == "false":
            cell.fill = PatternFill(fill_type="solid", fgColor="0F5132")
            cell.font = Font(color="FFFFFF", bold=True)

    def style_readiness(ws, row_idx, row_data, headers):
        label = (row_data.get("readiness_label") or "").strip()
        if label in readiness_styles:
            fill, font = readiness_styles[label]
            label_col = headers.index("readiness_label") + 1
            ws.cell(row=row_idx, column=label_col).fill = fill
            ws.cell(row=row_idx, column=label_col).font = font

        go_no_go_col = headers.index("go_no_go") + 1
        go_no_go = (row_data.get("go_no_go") or "").strip()
        go_cell = ws.cell(row=row_idx, column=go_no_go_col)
        if go_no_go == "NO GO":
            go_cell.fill = PatternFill(fill_type="solid", fgColor="7A0916")
            go_cell.font = Font(color="FFFFFF", bold=True)
        elif go_no_go == "GO":
            go_cell.fill = PatternFill(fill_type="solid", fgColor="005F3C")
            go_cell.font = Font(color="FFFFFF", bold=True)

    detail_headers = [
        "path",
        "source_url",
        "test_url",
        "score",
        "status",
        "note",
        "root_cause",
        "change_bucket",
        "fix_code",
        "owner",
    ]
    cluster_headers = [
        "failure_cluster",
        "test_site",
        "section",
        "signature",
        "failures_in_cluster",
        "section_failures",
        "sections_with_signature",
        "total_failures_with_signature",
        "systemic_breakage",
        "systemic_reason",
    ]
    readiness_headers = [
        "test_site",
        "section",
        "total_pages",
        "blocker_count",
        "non_blocker_count",
        "systemic_cluster_count",
        "readiness_score",
        "readiness_label",
        "go_no_go",
        "summary_reason",
    ]

    write_sheet("Detail", detail_headers, result_rows, style_detail, section_subheads=True)
    write_sheet("FailureClusters", cluster_headers, cluster_rows, style_cluster)
    write_sheet("ReleaseReadiness", readiness_headers, readiness_rows, style_readiness)

    wb.save(xlsx_path)
    return xlsx_path

# === 2. SITEMAP SCRAPER ===
def auto_generate_paths(base_url):
    sitemap_url = f"{base_url.rstrip('/')}/sitemap"
    print(f"--- [STEP 1] Scraping Sitemap: {sitemap_url} ---")
    try:
        response = requests.get(sitemap_url, timeout=15, **request_kwargs_for_url(sitemap_url))
        response.raise_for_status()
        cleaned_paths = []
        try:
            root = ET.fromstring(response.content)
            namespace = {'ns': 'http://www.sitemaps.org/schemas/sitemap/0.9'}
            urls = [loc.text for loc in root.findall('.//ns:loc', namespace)]
            for url in urls:
                cleaned_paths.append(urlparse(url).path or "/")
        except:
            links = re.findall(r'href=["\'](/?[\w\-\/]+)["\']', response.text)
            for link in links:
                if link.startswith("/") and not any(ext in link for ext in ['.pdf', '.jpg', '.png', '.css', '.docx']):
                    cleaned_paths.append(link)

        if cleaned_paths:
            final_list = sorted(list(set(cleaned_paths)))
            with open(PATHS_FILE, "w") as f:
                for path in final_list: f.write(f"{path}\n")
            print(f"[OK] Found {len(final_list)} paths.")
            return True
        return False
    except Exception as e:
        print(f"[WARN] Sitemap Scrape Failed: {e}")
        return False

# === 3. AUDIT ENGINE WITH REVISED SCORING ===
async def audit_single_path(browser_context, path, test_base):
    async with semaphore:
        page = await browser_context.new_page()
        safe_name = path.strip("/").replace("/", "_") or "home"
        source_url = f"{SOURCE_BASE}{path}" if SOURCE_BASE else ""
        test_url = join_test_url(test_base, path)
        row = {
            "path": path,
            "test_site": test_site_label(test_base),
            "source_url": source_url,
            "test_url": test_url,
            "score": 0,
            "status": "Pending",
            "note": "",
            "source_status": "",
            "test_status": "",
            "redirect": False,
            "root_cause": "",
        }
        
        try:
            test_url = join_test_url(test_base, path)
            source_url = f"{SOURCE_BASE}{path}" if SOURCE_BASE else None
            source_status, source_redirect = probe_url(source_url) if source_url else (None, False)
            test_status, test_redirect = probe_url(test_url)
            redirect = source_redirect or test_redirect

            row.update(
                {
                    "source_status": source_status if source_status is not None else "",
                    "test_status": test_status if test_status is not None else "",
                    "redirect": redirect,
                }
            )
            
            # 1. Check for Redirect
            response = await page.goto(test_url, timeout=60000)
            
            if response is None:
                row.update({
                    "status": "ERROR",
                    "note": "Failed to navigate to test URL"
                })
            elif response.url.rstrip("/") != test_url.rstrip("/") or redirect:
                row.update({
                    "status": "REDIRECT",
                    "note": "Redirect detected"
                })
            elif response.status >= 400:
                if source_status == 200 and normalize_path(path) in redirect_override_paths:
                    row.update({
                        "status": "REDIRECT",
                        "note": "Intentional redirect/replacement path",
                        "redirect": True,
                    })
                else:
                    row.update({
                        "status": "FAIL", 
                        "note": f"Status {response.status}"
                    })
            else:
                # 2. Content Comparison
                if source_url:
                    await page.goto(source_url, timeout=60000)
                    source_text = await page.inner_text("body")
                else:
                    source_text = ""
                
                await page.goto(test_url, timeout=60000)
                test_text = await page.inner_text("body")
                
                if source_text:
                    score = SequenceMatcher(None, source_text, test_text).ratio()
                else:
                    score = 1.0  # Default to PASS if no source comparison available
                row["score"] = round(score, 4)

                # --- REVISED SCORING LOGIC ---
                if score >= PASS_THRESHOLD:
                    row["status"] = "PASS"
                    row["note"] = "Match"
                elif score >= SOFT_PASS_THRESHOLD:
                    row["status"] = "SOFT PASS"
                    row["note"] = f"Minor Diff: {score:.1%} match"
                elif score > 0:
                    row["status"] = "REVIEW"
                    row["note"] = f"Diff: {score:.1%} match"
                    # Capture screenshots for manual review
                    await page.goto(source_url)
                    await page.screenshot(path=os.path.join(BASE_DIR, f"{safe_name}_OLD.png"), full_page=True)
                    await page.goto(test_url)
                    await page.screenshot(path=os.path.join(BASE_DIR, f"{safe_name}_NEW.png"), full_page=True)
                else:
                    row["status"] = "FAIL"
                    row["note"] = "Rendering Error/Empty Page"

            row["root_cause"] = determine_root_cause(
                row["status"],
                row["score"],
                source_status,
                test_status,
                redirect,
            )

        except Exception as e:
            row.update({"status": "ERROR", "note": str(e)[:100]})
            row["root_cause"] = determine_root_cause(
                row["status"],
                row["score"],
                row["source_status"] if isinstance(row["source_status"], int) else None,
                row["test_status"] if isinstance(row["test_status"], int) else None,
                bool(row["redirect"]),
            )
        
        await page.close()
        print(f"   [{row['status']}] {path}")
        return row

# === 4. CONTROLLER ===
async def run_audit():
    global TEST_BASE
    # If SOURCE_BASE is provided, scrape sitemap; otherwise use existing paths.txt
    if SOURCE_BASE:
        if not auto_generate_paths(SOURCE_BASE): return
    elif not os.path.exists(PATHS_FILE):
        print(f"[WARN] No --source provided and {PATHS_FILE} not found. Provide --source to auto-generate paths.")
        return

    # Load paths from file
    if not os.path.exists(PATHS_FILE):
        print(f"[WARN] {PATHS_FILE} not found after sitemap processing.")
        return
    
    with open(PATHS_FILE, "r") as f:
        paths = [line.strip() for line in f if line.strip()]

    test_scope = resolve_test_scope(args.test_scope)
    prompt_on_unhealthy = test_scope == "single"
    raw_test_bases = [TEST_BASE] if test_scope == "single" else discover_instance_test_bases(TEST_BASE)
    allow_labels, allow_urls = build_test_allowlist(args.test_allowlist, args.test_allowlist_file)
    if test_scope == "instance" and (allow_labels or allow_urls):
        pre_filter_count = len(raw_test_bases)
        raw_test_bases = filter_test_bases_by_allowlist(raw_test_bases, allow_labels, allow_urls)
        allow_source = []
        if args.test_allowlist.strip():
            allow_source.append(args.test_allowlist)
        if args.test_allowlist_file.strip():
            allow_source.append(f"file:{args.test_allowlist_file}")
        print(
            f"--- [STEP 2.1] Applied allowlist ({' | '.join(allow_source)}); "
            f"kept {len(raw_test_bases)} of {pre_filter_count} discovered targets ---"
        )
        if not raw_test_bases:
            print("[STOP] No test targets matched allowlist inputs. Check labels/URLs/file and retry.")
            return

    resolved_test_bases = []
    for base in raw_test_bases:
        normalized = normalize_site_base(base)
        if not normalized:
            continue

        print(f"--- [STEP 2] Validating Test Environment: {normalized} ---")
        try:
            test_check = requests.get(normalized, timeout=30, **request_kwargs_for_url(normalized))
            if test_check.status_code != 200:
                print(f"[STOP] Test server returned {test_check.status_code} for {normalized}.")
                if prompt_on_unhealthy:
                    if input("Continue with this target anyway? (y/n): ").lower() != 'y':
                        continue
                else:
                    continue
        except Exception as e:
            print(f"[WARN] Connection Error for {normalized}: {e}")
            continue

        corrected = maybe_correct_test_base(SOURCE_BASE, normalized, paths)
        if corrected not in resolved_test_bases:
            resolved_test_bases.append(corrected)

    if not resolved_test_bases:
        print("[STOP] No valid test targets available after validation.")
        return

    TEST_BASE = resolved_test_bases[0]
    print(f"--- [STEP 2.4] Test targets selected ({len(resolved_test_bases)}):")
    for base in resolved_test_bases:
        print(f"   - {base}")

    if args.max_paths and args.max_paths > 0:
        paths = paths[:args.max_paths]
        print(f"--- [STEP 2.5] Path limit enabled: auditing first {len(paths)} paths ---")

    async with async_playwright() as p:
        print(f"--- [STEP 3] Launching Audit (Concurrency: {args.max_tabs}) ---")
        browser = await p.chromium.launch(headless=True)
        ignore_https_errors = any(is_localhost_url(base) for base in resolved_test_bases)
        if ignore_https_errors:
            print("--- [STEP 3.1] Localhost target detected: ignoring local HTTPS certificate errors ---")
        context = await browser.new_context(
            user_agent="WSU-Auditor-v2",
            ignore_https_errors=ignore_https_errors,
        )

        tasks = [audit_single_path(context, path, test_base) for test_base in resolved_test_bases for path in paths]
        results = await asyncio.gather(*tasks)
        cluster_summary_rows = cluster_failure_patterns(results)
        release_readiness_rows = build_release_readiness(results)
        results = sort_audit_results(results)

        dated_report_name = f"{SAFE_SITE_NAME}_audit_report_{RUN_DATE}.csv"
        report_path = os.path.join(BASE_DIR, dated_report_name)
        legacy_report_path = os.path.join(BASE_DIR, "_audit_report.csv")
        with open(report_path, 'w', newline='') as f:
            writer = csv.DictWriter(
                f,
                fieldnames=[
                    "path",
                    "source_url",
                    "test_url",
                    "score",
                    "status",
                    "note",
                    "root_cause",
                    "change_bucket",
                    "fix_code",
                    "owner",
                ],
                extrasaction="ignore",
            )
            writer.writeheader()
            writer.writerows(results)

        cluster_summary_path = write_cluster_summary(
            BASE_DIR,
            SAFE_SITE_NAME,
            RUN_DATE,
            cluster_summary_rows,
        )
        release_readiness_path = write_release_readiness_summary(
            BASE_DIR,
            SAFE_SITE_NAME,
            RUN_DATE,
            release_readiness_rows,
        )
        html_report_path = write_accessible_html_report(
            BASE_DIR,
            SAFE_SITE_NAME,
            RUN_DATE,
            results,
            cluster_summary_rows,
            release_readiness_rows,
        )
        executive_report_path = write_executive_html_report(
            BASE_DIR,
            SAFE_SITE_NAME,
            RUN_DATE,
            results,
            release_readiness_rows,
        )
        xlsx_report_path = write_xlsx_report(
            BASE_DIR,
            SAFE_SITE_NAME,
            RUN_DATE,
            results,
            cluster_summary_rows,
            release_readiness_rows,
        )

        # Keep legacy filename for downstream scripts while using dated report as primary output.
        shutil.copyfile(report_path, legacy_report_path)

        await browser.close()
        print(f"\n[OK] Audit Complete! Results: {BASE_DIR}")
        print(f"   Primary report: {report_path}")
        print(f"   Legacy copy: {legacy_report_path}")
        print(f"   Cluster summary: {cluster_summary_path}")
        print(f"   Release readiness: {release_readiness_path}")
        print(f"   Accessible HTML report: {html_report_path}")
        print(f"   Executive HTML view: {executive_report_path}")
        print(f"   XLSX workbook: {xlsx_report_path}")

if __name__ == "__main__":
    asyncio.run(run_audit())